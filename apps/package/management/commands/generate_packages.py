import io
from django.core.files.base import File
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.db import transaction
from django.contrib.postgres.aggregates import StringAgg
from django.core.management.base import BaseCommand
from django.db.models import Sum, F, Q
from django.db import IntegrityError

from apps.publisher.models import Publisher
from apps.order.models import Order, OrderWindow, BookOrder
from apps.package.models import (
    PublisherPackage, PublisherPackageBook,
    SchoolPackage, SchoolPackageBook,
    InstitutionPackage, InstitutionPackageBook,
    CourierPackage
)
from apps.user.models import User
from apps.common.models import Municipality
# from apps.package.seed.incentive import INCENTIVE_BOOKS


class Command(BaseCommand):
    help = 'Generate publisher packages'

    def add_arguments(self, parser):
        parser.add_argument('order_window_id', type=int, help='order window id')

    def _format_unverified_users(self, unverified_user_qs):
        return '\n'.join(
            ['id = %s ---- full name = %s' % (
                user['created_by__id'], user['created_by__full_name']) for user in unverified_user_qs]
        )

    def _format_unverified_payments(self, unverified_payments_qs):
        return '\n'.join(
            ['id = %s ---- full name = %s' % (user['id'], user['paid_by__full_name']) for user in unverified_payments_qs]
        )

    def _format_mismatched_order_users(self, unverified_user_qs):
        return '\n'.join(
            ['id = %s ---- full name = %s' % (
                user['id'], user['full_name']) for user in unverified_user_qs]
        )

    def _generate_related_orders_export(
        self,
        package,
        related_book_orders,
        # total_school_book_quantity_list,
        for_school
    ):

        # def _append_or_update_incentive(incentive_list, book, internal_code):
        #     if incentive_list:
        #         for kitab_incentive in incentive_list:
        #             if (
        #                 book['book_name'] == kitab_incentive.get('book_name') and
        #                 book['publisher_name'] == kitab_incentive.get('publisher_name') and
        #                 book['internal_code'] == internal_code
        #             ):
        #                 # Increment quantity by one
        #                 kitab_incentive['quantity'] += 1
        #                 return kitab_incentive
        #     return incentive_list.append(book)

        filename = f'{package.publisher.name}book_orders.xlsx'
        wb = Workbook(write_only=True)
        ws1 = wb.create_sheet("Book orders")
        ws1.append([
            "Package Id", "Book Name", "Book Author/Authors", "Book Grade", "Book ISBN",
            "Book Edition", "Book Language", "Book Quantity", "Unit Book price",
            "Sub Total Book price"
        ])

        # To add sum formula at bottom
        book_grand_total = 0
        for book_order in related_book_orders:
            ws1.append([
                str(package.package_id),
                book_order['book__title'] if book_order['book__title'] else "",
                book_order['book_authors'] if book_order['book_authors'] else "",
                book_order['book__grade'] if book_order['book__grade'] else "",
                book_order['book__isbn'] if book_order['book__isbn'] else "",
                book_order['book__edition'] if book_order['book__edition'] else "",
                book_order['book__language'] if book_order['book__language'] else "",
                book_order['total_quantity'] if book_order['total_quantity'] else 0,
                book_order['book__price'] if book_order['total_quantity'] else 0,
                book_order['total_price'] if book_order['total_price'] else 0
            ])
            book_grand_total += book_order['total_quantity'] * book_order['book__price']

        if for_school:
            # XXX: Why is ws1 Grand total price here?
            ws1.append([
                "", "", "", "", "", "", "", "", "Grand Total Price", book_grand_total
            ])
            incentive_list = []
            ws2 = wb.create_sheet("Incentive Orders")
            ws2.append([
                "Book Name", "Unit Price", "Quantity", "Sub Total Price"
            ])

            # order_window = package.order_window
            # for total_school_book_quantity in total_school_book_quantity_list:
            #     # A school can can maximum 120 incentives
            #     if latest_order_window.enable_incentive and (
            #         total_school_book_quantity >= order_window.incentive_quantity_threshold
            #     ):
            #         incentive_quantity = min(
            #             total_school_book_quantity * order_window.incentive_multiplier,
            #             order_window.incentive_max,
            #         )

            #         for book in INCENTIVE_BOOKS[f'book_list_{incentive_quantity}']:
            #             _append_or_update_incentive(incentive_list, book, package.publisher.internal_code)

            # To add sum formula at bottom
            book_incentive_grand_total = 0
            for incentive in incentive_list:
                if incentive['internal_code'] == package.publisher.internal_code:
                    sub_total = incentive['price'] * incentive['quantity']
                    ws2.append(
                        [
                            incentive['book_name'], incentive['price'],
                            incentive['quantity'], sub_total
                        ]
                    )
                    book_incentive_grand_total += sub_total
            ws2.append([
                "", "", "Grand Total Price", book_incentive_grand_total
            ])

        fileToUpload = io.BytesIO(save_virtual_workbook(wb))
        package.orders_export_file.save(filename, File(fileToUpload))

    def _create_publisher_packages(self, latest_order_window, orders, for_school=False):
        # ------------------------------------------------------------------
        # Create publisher packages
        # ------------------------------------------------------------------

        # Incentive calculation should be done before filtering publishers
        # We calculate incentives based on school orders

        # total_school_book_quantity_list = orders.values('created_by').annotate(
        #     grand_total_quantity=Sum('book_order__quantity')
        # ).values_list('grand_total_quantity', flat=True)

        # Get unique publishers in order
        publisher_ids = orders.values_list('book_order__publisher', flat=True).distinct()

        publisher_package_count = 0

        # Create packages for each publishers
        for publisher_id in publisher_ids:
            publisher = Publisher.objects.get(id=publisher_id)
            related_orders = orders.filter(book_order__publisher=publisher)
            related_book_orders = BookOrder.objects.filter(
                order__in=related_orders, publisher=publisher
            ).values('book').annotate(
                total_quantity=Sum('quantity'),
                total_price=Sum(F('quantity') * F('price')),
                book_authors=StringAgg('book__authors__name', distinct=True, delimiter=", "),
                book__id=F('book__id'),
                book__title=F('book__title'),
                book__grade=F('book__grade'),
                book__language=F('book__language'),
                book__isbn=F('book__isbn'),
                book__price=F('book__price'),
                book__edition=F('book__edition'),
            )
            package = PublisherPackage.objects.create(
                status=PublisherPackage.Status.PENDING.value,
                publisher=publisher,
                order_window=latest_order_window,
                total_quantity=related_book_orders.aggregate(
                    grand_total_quantity=Sum('total_quantity'))['grand_total_quantity'],
                total_price=related_book_orders.aggregate(
                    grand_total_price=Sum('total_price'))['grand_total_price']
            )
            package.related_orders.set(related_orders)
            # Generate related orders export file
            self._generate_related_orders_export(
                package,
                related_book_orders,
                # total_school_book_quantity_list,
                for_school,
            )
            PublisherPackageBook.objects.bulk_create(
                [
                    PublisherPackageBook(
                        book_id=related_book_order['book__id'],
                        quantity=related_book_order['total_quantity'],
                        publisher_package=package
                    ) for related_book_order in related_book_orders
                ]
            )
            publisher_package_count += 1

        self.stdout.write(self.style.SUCCESS(f'{publisher_package_count} Publisher packages created.'))

    def _create_courier_packages_for_school(self, latest_order_window, orders):
        # ------------------------------------------------------------------
        # Create courier packages
        # ------------------------------------------------------------------
        courier_package_count = 0
        municipality_ids = orders.values_list(
            'created_by__school__municipality', flat=True
        ).distinct()
        for municipality_id in municipality_ids:
            courier_related_book_orders = orders.filter(
                created_by__school__municipality=municipality_id
            ).values(
                'created_by__school__municipality'
            ).annotate(
                total_quantity=Sum('book_order__quantity'),
                total_price=Sum('book_order__quantity') * F('book_order__price')
            ).values('total_quantity', 'total_price')

            courier_package = CourierPackage.objects.create(
                status=CourierPackage.Status.PENDING.value,
                order_window=latest_order_window,
                municipality=Municipality.objects.get(id=municipality_id),
                total_quantity=courier_related_book_orders.aggregate(
                    grand_total_quantity=Sum('total_quantity'))['grand_total_quantity'],
                total_price=courier_related_book_orders.aggregate(
                    grand_total_price=Sum('total_price'))['grand_total_price'],
                type=CourierPackage.Type.SCHOOL.value
            )
            if latest_order_window.enable_incentive and (
                courier_package.total_quantity >= latest_order_window.incentive_quantity_threshold
            ):
                courier_package.is_eligible_for_incentive = True
                courier_package.save()
            courier_package_count += 1
        self.stdout.write(self.style.SUCCESS(f'{courier_package_count} municipality/courier packages created.'))

    def _create_courier_packages_for_institution(self, latest_order_window, orders):
        # ------------------------------------------------------------------
        # Create courier packages
        # ------------------------------------------------------------------
        courier_package_count = 0
        municipality_ids = orders.values_list(
            'created_by__institution__municipality', flat=True
        ).distinct()
        for municipality_id in municipality_ids:
            courier_related_book_orders = orders.filter(
                created_by__institution__municipality=municipality_id
            ).values(
                'created_by__institution__municipality'
            ).annotate(
                total_quantity=Sum('book_order__quantity'),
                total_price=Sum(F('book_order__quantity') * F('book_order__price'))
            )

            CourierPackage.objects.create(
                status=CourierPackage.Status.PENDING.value,
                order_window=latest_order_window,
                municipality=Municipality.objects.get(id=municipality_id),
                total_quantity=courier_related_book_orders.aggregate(
                    grand_total_quantity=Sum('total_quantity'))['grand_total_quantity'],
                total_price=courier_related_book_orders.aggregate(
                    grand_total_price=Sum('total_price'))['grand_total_price'],
                is_eligible_for_incentive=False,
                type=CourierPackage.Type.INSTITUTION.value,
            )
            courier_package_count += 1
        self.stdout.write(self.style.SUCCESS(f'{courier_package_count} municipality/courier packages created.'))

    def _create_school_packages(self, latest_order_window, orders):
        # ------------------------------------------------------------------
        # Create school packages
        # ------------------------------------------------------------------

        # Get unique schools in order
        school_user_ids = orders.values_list('created_by', flat=True).distinct('created_by')
        school_package_count = 0

        # Create packages for each school
        for school_id in school_user_ids:
            school_user = User.objects.get(id=school_id)
            related_orders = orders.filter(created_by=school_user)
            related_book_orders = BookOrder.objects.filter(order__in=related_orders).values('book__created_by').annotate(
                total_quantity=Sum('quantity'),
                total_price=Sum(F('quantity') * F('price'))
            ).values('book', 'total_quantity', 'total_price')
            courier_package = CourierPackage.objects.get(
                order_window=latest_order_window, municipality=school_user.school.municipality
            )
            school_package = SchoolPackage.objects.create(
                status=SchoolPackage.Status.PENDING.value,
                school=school_user,
                order_window=latest_order_window,
                courier_package=courier_package,
                total_quantity=related_book_orders.aggregate(
                    grand_total_quantity=Sum('total_quantity'))['grand_total_quantity'],
                total_price=related_book_orders.aggregate(
                    grand_total_price=Sum('total_price'))['grand_total_price']
            )
            school_package.related_orders.set(related_orders)
            if latest_order_window.enable_incentive and (
                school_package.total_quantity >= latest_order_window.incentive_quantity_threshold
            ):
                school_package.is_eligible_for_incentive = True
                school_package.save()

                # Increment incentive
                PublisherPackage.objects.filter(related_orders__in=related_orders).distinct().update(
                    incentive=F('incentive') + 1
                )

            SchoolPackageBook.objects.bulk_create(
                [
                    SchoolPackageBook(
                        book_id=related_book_order['book'],
                        quantity=related_book_order['total_quantity'],
                        school_package=school_package,
                    ) for related_book_order in related_book_orders
                ]
            )
            school_package_count += 1

        self.stdout.write(self.style.SUCCESS(f'{school_package_count} School packages created.'))

    def _create_institution_packages(self, latest_order_window, orders):
        # ------------------------------------------------------------------
        # Create institution packages
        # ------------------------------------------------------------------

        # Get unique institution in order
        institution_user_ids = orders.values_list('created_by', flat=True).distinct('created_by')
        institution_package_count = 0

        # Create packages for each institution
        for institution_user_id in institution_user_ids:
            institution_user = User.objects.get(id=institution_user_id)
            related_orders = orders.filter(created_by=institution_user)
            related_book_orders = BookOrder.objects.filter(order__in=related_orders).values('book__created_by').annotate(
                total_quantity=Sum('quantity'),
                total_price=Sum(F('quantity') * F('price')),
                book=F('book')
            )
            courier_package = CourierPackage.objects.get(
                order_window=latest_order_window, municipality=institution_user.institution.municipality
            )
            institution_package = InstitutionPackage.objects.create(
                status=InstitutionPackage.Status.PENDING.value,
                institution=institution_user,
                order_window=latest_order_window,
                courier_package=courier_package,
                total_quantity=related_book_orders.aggregate(
                    grand_total_quantity=Sum('total_quantity'))['grand_total_quantity'],
                total_price=related_book_orders.aggregate(
                    grand_total_price=Sum('total_price'))['grand_total_price']
            )
            institution_package.related_orders.set(related_orders)
            InstitutionPackageBook.objects.bulk_create(
                [
                    InstitutionPackageBook(
                        book_id=related_book_order['book'],
                        quantity=related_book_order['total_quantity'],
                        school_package=institution_package,
                    ) for related_book_order in related_book_orders
                ]
            )
            institution_package_count += 1

        self.stdout.write(self.style.SUCCESS(f'{institution_package_count} Institution packages created.'))

    def _generate_school_packages(self, latest_order_window, orders):
        self._create_publisher_packages(latest_order_window, orders, for_school=True)
        self._create_courier_packages_for_school(latest_order_window, orders)
        self._create_school_packages(latest_order_window, orders)

    def _generate_institution_packages(self, latest_order_window, orders):
        self._create_publisher_packages(latest_order_window, orders, for_school=False)
        self._create_courier_packages_for_institution(latest_order_window, orders)
        self._create_institution_packages(latest_order_window, orders)

    @transaction.atomic
    def handle(self, *args, **options):
        order_window_id = options['order_window_id']
        # Check if order window exists
        try:
            latest_order_window = OrderWindow.objects.get(id=order_window_id)
        except OrderWindow.DoesNotExist:
            self.stdout.write(self.style.ERROR('Invalid order window id supplied.'))
            return

        # Get orders belongs to order window
        orders = Order.objects.filter(
            book_order__publisher__isnull=False,
            assigned_order_window__id=latest_order_window.pk,
            status=Order.Status.PENDING.value,
            created_by__is_deactivated=False
        )
        user_ids = orders.values_list('created_by__id', flat=True)
        mismatched_order_users = User.objects.filter(
            id__in=user_ids, is_deactivated=False).annotate(**User.annotate_mismatch_order_statements()).filter(
            outstanding_balance__lt=0
        )
        if mismatched_order_users.exists():
            formatted_mismatched_order_users = mismatched_order_users.values('id', 'full_name')
            formated_unverified_users = self._format_mismatched_order_users(formatted_mismatched_order_users)
            self.stdout.write(self.style.ERROR(
                'Mismatched orders exists please fix those \n'
                f'{formated_unverified_users}'
            ))
            return

        # Check if unverified users exists and their profile, we need municipality
        # to generate courier packages
        if latest_order_window.type == OrderWindow.OrderWindowType.SCHOOL.value:
            unverified_users_qs = orders.filter(
                Q(created_by__is_verified=False) | Q(created_by__school__isnull=True)
            ).distinct()
        elif latest_order_window.type == OrderWindow.OrderWindowType.INSTITUTION.value:
            unverified_users_qs = orders.filter(
                Q(created_by__is_verified=False) | Q(created_by__institution__isnull=True)
            ).distinct()
        else:
            raise Exception(f'Unknown order window type: {latest_order_window.type}')

        if unverified_users_qs.exists():
            unverified_users = unverified_users_qs.values('created_by__id', 'created_by__full_name')
            formated_unverified_users = self._format_unverified_users(unverified_users)
            self.stdout.write(self.style.ERROR(
                'Following users are not verified or school profile is not attached \n'
                f'{formated_unverified_users}'
            ))
            return

        # Check if packages are already created
        if (
            PublisherPackage.objects.filter(order_window=latest_order_window).exists() or
            SchoolPackage.objects.filter(order_window=latest_order_window).exists() or
            CourierPackage.objects.filter(order_window=latest_order_window).exists() or
            InstitutionPackage.objects.filter(order_window=latest_order_window).exists()
        ):
            f'Packages for order window {latest_order_window} are already created.'

        try:
            if latest_order_window.type == OrderWindow.OrderWindowType.SCHOOL:
                self._generate_school_packages(latest_order_window, orders)
            elif latest_order_window.type == OrderWindow.OrderWindowType.INSTITUTION:
                self._generate_institution_packages(latest_order_window, orders)
            else:
                self.stdout.write(self.style.ERROR('Invalid order window'))
        except IntegrityError:
            self.stdout.write(self.style.ERROR(
                f'Packages for order window {latest_order_window} are already created.'
            ))
            return
